from openpyxl import load_workbook
import os
import re
import pandas as pd


class ExcelReader():

    def __init__(self, path = "input\\", filename = str) -> None:
        self.path = path
        self.filename = filename
        self.workbook = load_workbook(filename= path +filename, data_only=True, keep_links=False)

    def __repr__(self) -> str:
        myString = "Instance de la classe ExcelFile\n- Nom du fichier : {}".format(self.filename)
        return myString

    def get_sheet(self):
        return self.workbook.sheetnames

    def get_table(self, sheet_name, table_name):
        a = self.workbook[sheet_name].tables[table_name].ref.split(':')

        for i in range(len(a)):
            a[i] = re.split('(\d+)', a[i])[:-1]

        df = pd.read_excel(self.path + self.filename, sheet_name=sheet_name, skiprows=int(a[0][1]) - 1, nrows=int(a[1][1])-int(a[0][1]),
                      usecols=a[0][0] + ":" + a[1][0])
        return df

    def get_value(self, value_name):
        temp = self.workbook.defined_names[value_name].attr_text.split('$', 1)
        sheet = temp[0][:-1]
        cell = temp[1]
        return self.workbook[sheet][cell].value


    def get_range(self, sheet_name, cols, skiprows, nrows, names):
        """
        Parameters
        ----------
        - sheet_name (str) \t Name of the sheet
        - cols (str) \t\t Columns used (ex : 'A:N' or 'A, C, F:T")
        - skiprows (int) \t Number of rows to skip
        - nrows (int) \t\t Length of the Excel tab

        Return
        ------
        * (df) \t Excel range in a pandas dataframe.

        Example
        -------
        >>> objet.get_range(sheet_name="My Sheet", cols='F:M', skiprows=5, nrows=10)

        """
        file_location = self.path + self.filename
        return pd.read_excel(file_location,sheet_name=sheet_name, usecols=cols, skiprows=skiprows, nrows=nrows, names = names)
